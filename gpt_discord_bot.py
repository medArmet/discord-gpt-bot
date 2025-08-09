import os
import re
import io
import csv
import json
import math
import asyncio
import aiohttp
import discord
from typing import List, Tuple
from openai import OpenAI

# Optional parsers (installed via requirements.txt)
import pdfplumber
from docx import Document as DocxDocument
import openpyxl

# =======================
# Config
# =======================
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
DISCORD_TOKEN  = os.getenv("DISCORD_BOT_TOKEN")
PRIMARY_MODEL  = "gpt-5"
FALLBACK_MODEL = "gpt-5-mini"   # used only if primary returns empty

if not OPENAI_API_KEY:
    raise RuntimeError("OPENAI_API_KEY is not set")
if not DISCORD_TOKEN:
    raise RuntimeError("DISCORD_BOT_TOKEN is not set")

client_openai = OpenAI(api_key=OPENAI_API_KEY)

intents = discord.Intents.default()
intents.message_content = True
client = discord.Client(intents=intents)

# Accept "!gpt" or "gpt" (case-insensitive) at the start
PREFIX_RE = re.compile(r"^\s*!?gpt\b", re.IGNORECASE)

# Token/size guards (rough char budgets; adjust as you like)
MAX_TEXT_PER_FILE = 60_000     # chars per file after parsing (pre-truncation)
MAX_TEXT_ALL_FILES = 120_000   # combined chars across all files
CSV_PREVIEW_ROWS = 60          # more generous preview by default
XLSX_PREVIEW_ROWS = 40
XLSX_PREVIEW_COLS = 25


# =======================
# Utilities
# =======================
async def download_bytes(url: str) -> bytes:
    async with aiohttp.ClientSession() as session:
        async with session.get(url) as resp:
            resp.raise_for_status()
            return await resp.read()

def safe_truncate(text: str, limit: int) -> str:
    if len(text) <= limit:
        return text
    return text[:limit] + "\n...[truncated]..."

def csv_to_preview(data: bytes, limit_rows: int = CSV_PREVIEW_ROWS) -> str:
    # Try utf-8 → latin-1
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1", errors="ignore")
    # Strip BOM
    if text and text[0] == "\ufeff":
        text = text[1:]

    buf = io.StringIO(text)
    reader = csv.reader(buf)
    rows = []
    for i, row in enumerate(reader):
        if i >= limit_rows:
            rows.append(["...[truncated rows]..."])
            break
        rows.append(row)

    out = io.StringIO()
    csv.writer(out).writerows(rows)
    return out.getvalue()

def json_to_preview(data: bytes, limit_chars: int = MAX_TEXT_PER_FILE) -> str:
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        text = data.decode("latin-1", errors="ignore")
    try:
        obj = json.loads(text)
        pretty = json.dumps(obj, indent=2, ensure_ascii=False)
    except Exception:
        pretty = text
    return safe_truncate(pretty, limit_chars)

def txt_like_to_text(data: bytes, limit_chars: int = MAX_TEXT_PER_FILE) -> str:
    try:
        text = data.decode("utf-8", errors="ignore")
    except Exception:
        try:
            text = data.decode("latin-1", errors="ignore")
        except Exception:
            text = ""
    return safe_truncate(text, limit_chars)

def pdf_to_text(data: bytes, limit_chars: int = MAX_TEXT_PER_FILE) -> str:
    text = ""
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            piece = page.extract_text() or ""
            text += piece + "\n"
            if len(text) > limit_chars:
                break
    return safe_truncate(text, limit_chars)

def docx_to_text(data: bytes, limit_chars: int = MAX_TEXT_PER_FILE) -> str:
    doc = DocxDocument(io.BytesIO(data))
    parts = []
    for p in doc.paragraphs:
        parts.append(p.text)
        if sum(len(x) for x in parts) > limit_chars:
            break
    text = "\n".join(parts)
    return safe_truncate(text, limit_chars)

def xlsx_to_preview(data: bytes,
                    limit_rows: int = XLSX_PREVIEW_ROWS,
                    limit_cols: int = XLSX_PREVIEW_COLS,
                    limit_chars: int = MAX_TEXT_PER_FILE) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = []
    row_count = 0
    for r in sheet.iter_rows(min_row=1,
                             max_row=limit_rows,
                             max_col=limit_cols,
                             values_only=True):
        row_count += 1
        vals = ["" if v is None else str(v) for v in r]
        rows.append(vals)
    # Convert to CSV-like preview
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerows(rows)
    preview = out.getvalue()
    if sheet.max_row > limit_rows:
        preview += "\n...[truncated rows]...\n"
    return safe_truncate(preview, limit_chars)

def ext(fname: str) -> str:
    return os.path.splitext(fname)[1].lower()

def extract_text_from_any_file(filename: str, content_type: str, data: bytes) -> Tuple[str, str]:
    """
    Returns (kind, extracted_text_or_preview)
    kind is a short descriptor shown to the model (e.g., "CSV preview", "PDF text", "DOCX text", "Text", "JSON", "XLSX preview", "Binary").
    """
    ctype = (content_type or "").lower()
    extension = ext(filename)

    # Images: we don't extract text here—handled as image_url upstream
    if ctype.startswith("image/") or extension in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"]:
        return "Image", ""

    # CSV
    if ctype.endswith("/csv") or extension == ".csv":
        return "CSV preview", csv_to_preview(data)

    # JSON
    if ctype.endswith("/json") or extension == ".json":
        return "JSON", json_to_preview(data)

    # PDF
    if ctype == "application/pdf" or extension == ".pdf":
        try:
            return "PDF text", pdf_to_text(data)
        except Exception:
            # fall back to binary notice
            return "Binary", ""

    # DOCX
    if extension == ".docx" or ctype in ("application/vnd.openxmlformats-officedocument.wordprocessingml.document",):
        try:
            return "DOCX text", docx_to_text(data)
        except Exception:
            return "Binary", ""

    # XLSX
    if extension == ".xlsx" or ctype in ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",):
        try:
            return "XLSX preview", xlsx_to_preview(data)
        except Exception:
            return "Binary", ""

    # Plain-ish text (md, txt, html, csv mislabeled)
    if extension in [".txt", ".md", ".html", ".htm", ".log", ".tsv", ".yaml", ".yml", ".ini"]:
        return "Text", txt_like_to_text(data)

    # Octet-stream or unknown → try as text, otherwise Binary
    try_text = txt_like_to_text(data)
    if try_text.strip():
        return "Text", try_text

    return "Binary", ""  # let the model know we couldn't decode


def extract_text_from_attachments(attachments: List[discord.Attachment]) -> Tuple[List[dict], int]:
    """
    For each attachment, produce one or more content blocks for the chat message.
    Returns (blocks, total_chars_used).
    """
    blocks: List[dict] = []
    total_chars = 0

    for a in attachments:
        filename = a.filename or "file"
        ctype = a.content_type or ""
        # Images: send as image_url (model can fetch URL directly)
        if (ctype.lower().startswith("image/")) or (ext(filename) in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"]):
            blocks.append({"type": "image_url", "image_url": {"url": a.url}})
            continue

        # Download for parsing
        try:
            file_bytes = asyncio.get_event_loop().run_until_complete(download_bytes(a.url))
        except RuntimeError:
            # if already in an event loop (we are), use await path (caller should pass data ideally)
            # but for simplicity, re-download synchronously in a nested loop via helper:
            pass

        file_bytes = None
        # Proper async download in this scope:
        # (We can't call await here; instead, download in caller. So:
        # We'll download above in the caller and pass raw bytes in.)
        # To keep this helper simple, we'll just skip here; real download happens in on_message.
        # (This helper is not used directly; see on_message where download happens.)

    # We don't actually use this helper. Left here for reference.
    return blocks, total_chars


def build_user_blocks_for_files(files_meta: List[Tuple[str, str, bytes]]) -> Tuple[List[dict], int]:
    """
    Takes a list of (filename, content_type, bytes) for all non-image files and returns content blocks.
    """
    blocks: List[dict] = []
    total_chars = 0

    for (filename, ctype, data) in files_meta:
        kind, text = extract_text_from_any_file(filename, ctype, data)

        if kind == "Image":
            # This path shouldn't occur here (images handled elsewhere), but guard anyway
            blocks.append({"type": "text", "text": f"Image file '{filename}' attached. (Processed separately.)"})
            continue

        if text:
            # enforce global cap
            remaining = MAX_TEXT_ALL_FILES - total_chars
            if remaining <= 0:
                blocks.append({"type": "text", "text": f"Additional file '{filename}' omitted due to size limits."})
                continue

            snippet = text if len(text) <= remaining else text[:remaining] + "\n...[truncated due to overall limit]..."
            total_chars += len(snippet)

            blocks.append({
                "type": "text",
                "text": f"{kind} from '{filename}':\n{snippet}"
            })
        else:
            # No text extracted (binary/unknown)
            size_kb = "unknown"
            try:
                size_kb = f"{math.ceil(len(data)/1024)} KB"
            except Exception:
                pass
            blocks.append({
                "type": "text",
                "text": (
                    f"File '{filename}' (type: {ctype or 'unknown'}, size: {size_kb}) "
                    f"could not be decoded as text. Suggest how to process it."
                )
            })

    return blocks, total_chars


def extract_text_from_completion(completion) -> str:
    try:
        msg = completion.choices[0].message
        if msg and getattr(msg, "content", None):
            return msg.content.strip()
    except Exception:
        pass
    return ""


async def call_openai_chat(model: str, messages: list, max_tokens: int = 1200):
    """Run blocking API call in a thread to avoid blocking Discord heartbeats."""
    def _do():
        return client_openai.chat.completions.create(
            model=model,
            messages=messages,
            max_completion_tokens=max_tokens,  # GPT-5 requires this name
        )
    return await asyncio.to_thread(_do)


def split_prefix_and_prompt(content: str) -> Tuple[bool, str]:
    m = PREFIX_RE.match(content)
    if not m:
        return False, content
    rest = content[m.end():].lstrip()
    return True, rest


# =======================
# Discord Events
# =======================
@client.event
async def on_ready():
    print(f"✅ Logged in as {client.user}")


@client.event
async def on_message(message: discord.Message):
    # Only handle messages that start with gpt / !gpt
    is_cmd, prompt = split_prefix_and_prompt(message.content)
    if message.author.bot or not is_cmd:
        return

    attachments = message.attachments or []
    print(f"DEBUG received cmd from {message.author} | msg_id={message.id} | attachments={len(attachments)}")
    if attachments:
        print("DEBUG attachment names:", [f"{a.filename} ({a.content_type})" for a in attachments])

    if not prompt and not attachments:
        await message.channel.send("❗ Please provide text or attach a file/image.")
        return

    thinking = await message.channel.send("⏳ Thinking...")

    try:
        # Build user content
        user_blocks: List[dict] = []

        # Add prompt text if any
        if prompt:
            user_blocks.append({"type": "text", "text": prompt})

        # Step 1: add image URLs directly
        non_image_files: List[Tuple[str, str, bytes]] = []
        for a in attachments:
            filename = a.filename or "file"
            ctype = (a.content_type or "").lower()
            if ctype.startswith("image/") or ext(filename) in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"]:
                user_blocks.append({"type": "image_url", "image_url": {"url": a.url}})
            else:
                # Download bytes for parsing
                file_bytes = await download_bytes(a.url)
                non_image_files.append((filename, ctype, file_bytes))

        # Step 2: parse non-image files into text blocks (PDF, DOCX, CSV, JSON, XLSX, TXT, etc.)
        file_blocks, _ = build_user_blocks_for_files(non_image_files)
        user_blocks.extend(file_blocks)

        # If user only sent files, ensure we nudge for plain-text output
        if not prompt:
            user_blocks.append({"type": "text", "text": "Please respond in plain text with a concise, well-structured analysis."})

        # Compose messages
        system_prompt = (
            "You are a capable analyst. Always respond in plain text (no markdown tables unless asked). "
            "Include: quick summary, key findings/patterns, and 3–5 actionable recommendations. "
            "If input includes images, use them as additional context (but do not hallucinate text from unreadable images)."
        )
        messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user",   "content": user_blocks if user_blocks else (prompt or "Please answer in plain text.")},
        ]

        # Log what we send
        print(f"DEBUG OpenAI call → model: {PRIMARY_MODEL}, discord_msg_id: {message.id}")
        # Log a concise preview
        brief = []
        for b in user_blocks:
            if b.get("type") == "image_url":
                brief.append({"type": "image_url", "url": b["image_url"].get("url", "")})
            else:
                t = b.get("text", "")
                brief.append({"type": "text", "text": (t[:140] + "…") if len(t) > 140 else t})
        print("DEBUG user content blocks:", brief[:6], ("...(+more)" if len(brief) > 6 else ""))

        # Primary call
        completion = await call_openai_chat(PRIMARY_MODEL, messages, max_tokens=1500)
        reply = extract_text_from_completion(completion)

        # Fallbacks if empty
        if not reply:
            print("DEBUG: Empty reply from primary. Retrying with text-only prompt on fallback model.")
            # Collapse into a simple text prompt for robustness
            plain_text_prompt = "Please analyze the provided content and respond in plain text with summary, key findings, and 3-5 action items."
            retry_messages = [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": plain_text_prompt},
            ]
            completion = await call_openai_chat(FALLBACK_MODEL, retry_messages, max_tokens=800)
            reply = extract_text_from_completion(completion)

        if not reply:
            reply = "⚠️ I couldn't generate a response."

        await thinking.edit(content=reply[:1900])  # Discord limit ~2000

    except Exception as e:
        await thinking.edit(content=f"❌ Error: {e}")


if __name__ == "__main__":
    client.run(DISCORD_TOKEN)
